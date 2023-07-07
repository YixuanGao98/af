import torch.utils.data as data
from PIL import Image
import os
import os.path
import scipy.io
import numpy as np
import csv
from openpyxl import load_workbook


class LIVEFolder(data.Dataset):

    def __init__(self, root, index, transform, patch_num):

        self.root = root
        # self.loader = loader

        self.refpath = os.path.join(self.root, 'refimgs')
        self.refname = getFileName( self.refpath,'.bmp')

        # self.refnames_all = []
        self.imgname=[]
        self.labels = []
        self.csv_file = os.path.join(self.root, 'LIVEhist_new.txt')
        with open(self.csv_file) as f:
            reader = f.readlines()
            for i, line in enumerate(reader):
                token = line.split("\t")
                token[0]=eval(token[0]) #LIVE去除字符串两端的引号
                self.imgname.append(token[0])
                values = np.array(token[11], dtype='float32')
                # values /= values.sum()
                self.labels.append(values)



        refnames_all = scipy.io.loadmat(os.path.join(self.root, 'refnames_all.mat'))
        self.refnames_all = refnames_all['refnames_all']

        self.dmos = scipy.io.loadmat(os.path.join(self.root, 'dmos_realigned.mat'))
        # self.labels = self.dmos['dmos_new'].astype(np.float32)   
        #self.labels = self.labels.tolist()[0]
        self.orgs = self.dmos['orgs']


        sample = []

        for i in range(0, len(index)):
            train_sel = (self.refname[index[i]] == self.refnames_all)
            train_sel = train_sel * ~self.orgs.astype(np.bool_)
            train_sel1 = np.where(train_sel == True)
            train_sel = train_sel1[1].tolist()
            for j, item in enumerate(train_sel):
                for aug in range(patch_num):
                #     sample.append((imgpath[item], labels[0][item]))
                    sample.append((os.path.join(self.root,'allimg', self.imgname[item]), self.labels[item]))
        self.samples = sample
        self.transform = transform

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = pil_loader(path)
        if self.transform is not None:
            sample = self.transform(sample)

        return sample, target

    def __len__(self):
        length = len(self.samples)
        return length

    def getDistortionTypeFileName(self, path, num):
        filename = []
        index = 1
        for i in range(0, num):
            name = '%s%s%s' % ('img', str(index), '.bmp')
            filename.append(os.path.join(path, name))
            index = index + 1
        return filename


class LIVEChallengeFolder(data.Dataset):

    def __init__(self, root, index, transform, patch_num):

        # imgpath = scipy.io.loadmat(os.path.join(root, 'Data', 'AllImages_release.mat'))
        # imgpath = imgpath['AllImages_release']
        # imgpath = imgpath[7:1169]
        # mos = scipy.io.loadmat(os.path.join(root, 'Data', 'AllMOS_release.mat'))
        # labels = mos['AllMOS_release'].astype(np.float32)
        # labels = labels[0][7:1169]

        self.root = root
        # self.loader = loader

        self.imgname = []
        self.labels = []
        self.csv_file = os.path.join(self.root, 'CLIVEhist.txt')
        with open(self.csv_file) as f:
            reader = f.readlines()
            for i, line in enumerate(reader):
                token = line.split("\t")
                token[0]=eval(token[0]) #LIVE去除字符串两端的引号
                self.imgname.append(token[0])
                values = np.array(token[11], dtype='float32')
                # values /= values.sum()
                self.labels.append(values)
        

        sample = []
        for i, item in enumerate(index):
            for aug in range(patch_num):
                sample.append((os.path.join(self.root,'Images', self.imgname[item]), self.labels[item]))

        self.samples = sample
        self.transform = transform

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = pil_loader(path)
        sample = self.transform(sample)
        return sample, target

    def __len__(self):
        length = len(self.samples)
        return length


class CSIQFolder(data.Dataset):



    def __init__(self, root, index, transform, patch_num):
        refpath = os.path.join(root, 'src_imgs')
        refname = getFileName(refpath,'.png')
   
   
   
        imgname = []
        mos_all = []
        refnames_all = []
        csv_file = os.path.join(root, 'CSIQhist.txt')
        with open(csv_file) as f:
            reader = f.readlines()
            for i, line in enumerate(reader):
                token = line.split("\t")
                token[0]=eval(token[0]) #LIVE去除字符串两端的引号
                imgname.append(token[0])
                values = np.array(token[11], dtype='float32')
                # values /= values.sum()
                mos_all.append(values)

                ref_temp = token[0].split(".")
                refnames_all.append(ref_temp[0] + '.' + ref_temp[-1])

        refnames_all = np.array(refnames_all)

        sample = []
        for i, item in enumerate(index):
            train_sel = (refname[index[i]] == refnames_all)
            train_sel = np.where(train_sel == True)
            train_sel = train_sel[0].tolist()
            for aug in range(patch_num):
                sample.append((os.path.join(root, 'dst_imgs_all', imgname[item]), mos_all[item]))
        self.samples = sample
        self.transform = transform

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = pil_loader(path)
        sample = self.transform(sample)

        return sample, target

    def __len__(self):
        length = len(self.samples)
        return length


class Koniq_10kFolder(data.Dataset):

    def __init__(self, root, index, transform, patch_num):
        imgname = []
        mos_all = []
        csv_file = os.path.join(root, 'koniq10k_scores_and_distributions.txt')
        with open(csv_file) as f:
            reader = f.readlines()
            for i, line in enumerate(reader):
                token = line.split("\t")
                # token[0]=eval(token[0]) #LIVE去除字符串两端的引号
                imgname.append(token[0])
                values = np.array(token[7], dtype='float32')
                # values /= values.sum()
                mos_all.append(values)

        sample = []
        for i, item in enumerate(index):
            for aug in range(patch_num):
                sample.append((os.path.join(root, 'koniq10k_1024x768/1024x768', imgname[item]), mos_all[item]))

        self.samples = sample
        self.transform = transform

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = pil_loader(path)
        sample = self.transform(sample)
        return sample, target

    def __len__(self):
        length = len(self.samples)
        return length


class BIDFolderAug(data.Dataset):

    def __init__(self, root, index, transform, patch_num):

        imgname = []
        mos_all = []

        xls_file = os.path.join(root, 'DatabaseGrades.xlsx')
        workbook = load_workbook(xls_file)
        booksheet = workbook.active
        rows = booksheet.rows
        count = 1
        for row in rows:
            count += 1
            img_num = (booksheet.cell(row=count, column=1).value)
            img_name = "DatabaseImage%04d.JPG" % (img_num)
            imgname.append(img_name)
            mos = (booksheet.cell(row=count, column=2).value)
            mos = np.array(mos)
            mos = mos.astype(np.float32)
            mos_all.append(mos)
            if count == 587:
                break

        sample = []
        for i, item in enumerate(index):
            for aug in range(patch_num):
                sample.append((os.path.join(root, imgname[item]), mos_all[item]))

        self.samples = sample
        self.transform = transform

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = pil_loader(path)
        sample = self.transform(sample)
        return sample, target

    def __len__(self):
        length = len(self.samples)
        return length


class TIDFolder(data.Dataset):

    def __init__(self, root, index, transform, patch_num):
        refpath = os.path.join(root, 'reference_images')
        refname = getTIDFileName(refpath,'.bmp.BMP')
        
        imgname = []
        mos_all = []
        refnames_all = []
        csv_file = os.path.join(root, 'TIDhist.txt')
        with open(csv_file) as f:
            reader = f.readlines()
            for i, line in enumerate(reader):
                token = line.split("\t")
                token[0]=eval(token[0]) #LIVE去除字符串两端的引号
                imgname.append(token[0])
                values = np.array(token[11], dtype='float32')
                # values /= values.sum()
                mos_all.append(values)
                ref_temp = token[0].split("_")
                refnames_all.append(ref_temp[0][1:])
                

        sample = []
        for i, item in enumerate(index):
            # print(refname[index[i]])
            train_sel = (refname[index[i]] == refnames_all)
            train_sel = np.where(train_sel == True)
            train_sel = train_sel[0].tolist()
            for j, item in enumerate(train_sel):
                for aug in range(patch_num):
                    sample.append((os.path.join(root, 'distorted_images', self.imgname[item]), self.histlabels[item]))
        self.samples = sample
        self.transform = transform
    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = pil_loader(path)
        sample = self.transform(sample)
        return sample, target

    def __len__(self):
        length = len(self.samples)
        return length

class LIVEMDFolder(data.Dataset):

    def __init__(self, root, index, transform, patch_num):
        
        # self.root = root
        # self.loader = loader
        self.refpath = os.path.join(root, 'refimgs')
        self.refname = getFileName( self.refpath,'.bmp')
        
        imgname = []
        refnames_all = []
        mos_all = []
        csv_file = os.path.join(root, 'LIVEMDhist.txt')
        with open(csv_file) as f:
            reader = f.readlines()
            for i, line in enumerate(reader):
                token = line.split("\t")
                token[0]=eval(token[0]) #LIVE去除字符串两端的引号
                imgname.append(token[0])
                values = np.array(token[11], dtype='float32')
                # values /= values.sum()
                mos_all.append(values)
        refnames_all = scipy.io.loadmat(os.path.join(root, 'refnames_all.mat'))
        self.refnames_all = refnames_all['refnames_all']



        sample = []
        for i, item in enumerate(index):
            train_sel = (self.refname[index[i]] == self.refnames_all)
            train_sel = np.where(train_sel == True)
            train_sel = train_sel[0].tolist()
            
            for aug in range(patch_num):
                sample.append((os.path.join(root, 'allimage', imgname[item]), mos_all[item]))
        self.samples = sample
        self.transform = transform
    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = pil_loader(path)
        sample = self.transform(sample)
        return sample, target

    def __len__(self):
        length = len(self.samples)
        return length



class CID2013Folder(data.Dataset):

    def __init__(self, root, index, transform, patch_num):
        
        self.root = root
        # self.loader = loader

        self.imgname = []
        self.mos=[]
        self.csv_file = '/home/gaoyixuan/imagehist3/bin5/CID2013hist_5.txt'
        with open(self.csv_file) as f:
            reader = f.readlines()
            for i, line in enumerate(reader):
                token = line.split("\t")
                token[0]=eval(token[0]) #LIVE去除字符串两端的引号
                self.imgname.append(token[0]+'.jpg')
                mos=eval(token[6])
                self.mos.append(mos)
        

        sample = []
        

        for i, item in enumerate(index):
            for aug in range(patch_num):
                sample.append((os.path.join(self.root,'CIDimg', self.imgname[item]), self.mos[item]))
            # sample.append((self.imgpath[item],self.labels[0][item]))
        self.samples = sample
        self.transform = transform
        
    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = pil_loader(path)
        sample = self.transform(sample)
        return sample, target

    def __len__(self):
        length = len(self.samples)
        return length



class VCLFERFolder(data.Dataset):

    def __init__(self, root, index, transform, patch_num):
        self.root = root

        refpath = os.path.join(root, 'Ref_Images')
        refname = getFileName(refpath,'.bmp')



        self.histlabels = []
        self.imgname=[]
        refnames_all = []
        self.csv_file = '/home/gaoyixuan/imagehist3/bin5/VCLFERhist.txt'
        # self.csv_file ='/home/gyx/DATA/imagehist/CSIQ/CSIQhist.txt'
        self.mos=[]
        self.std=[]
        with open(self.csv_file) as f:
            reader = f.readlines()
            for i, line in enumerate(reader):
                token = line.split("\t")
                token[0]=eval(token[0]) #LIVE去除字符串两端的引号
                token0 = token[0].split("/")
                token1=token0[2] #LIVE去除字符串两端的引号
                self.imgname.append(token1)
                values = np.array(token[1:6], dtype='float32')
                values /= values.sum()
                self.histlabels.append(values)

                ref_temp = token1.split(".")
                ref=ref_temp[0]
                refnames_all.append(ref[0:6] + '.bmp' )
                mos=eval(token[6])
                self.mos.append(mos)
                std=eval(token[7])
                self.std.append(std)

        # moslabels = np.array(target).astype(np.float32)
        refnames_all = np.array(refnames_all)

        sample = []

        for i, item in enumerate(index):
            train_sel = (refname[index[i]] == refnames_all)
            train_sel = np.where(train_sel == True)
            train_sel = train_sel[0].tolist()
            for j, item in enumerate(train_sel):
                # for aug in range(patch_num):
                sample.append((os.path.join(root, 'vcl_fer', self.imgname[item]), self.mos[item]))
        self.samples = sample
        self.transform = transform

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = pil_loader(path)
        sample = self.transform(sample)
        return sample, target

    def __len__(self):
        length = len(self.samples)
        return length




class SPAQFolder:
    def __init__(self, root, index, transform, patch_num):
    

        self.root = root

        self.imgname = []
        self.labels = []
        self.mos=[]
        self.std=[]
        self.csv_file = '/home/gaoyixuan/imagehist3/bin5/SPAQhist.txt'
        with open(self.csv_file) as f:
            reader = f.readlines()
            for i, line in enumerate(reader):
                token = line.split("\t")
                token[0]=eval(token[0]) #LIVE去除字符串两端的引号
                self.imgname.append(token[0])
                values = np.array(token[1:6], dtype='float32')
                values /= values.sum()
                self.labels.append(values)
                mos=eval(token[6])
                self.mos.append(mos)
                std=eval(token[7])
                self.std.append(std)

        sample = []
        
        for i, item in enumerate(index):
            sample.append((os.path.join(self.root, 'TestImage', self.imgname[item]), self.mos[item]))

        self.samples = sample    
        self.transform = transform


    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = pil_loader(path)
        sample = self.transform(sample)
        return sample, target

    def __len__(self):
        length = len(self.samples)
        return length




def getFileName(path, suffix):
    filename = []
    f_list = os.listdir(path)
    for i in f_list:
        if os.path.splitext(i)[1] == suffix:
            filename.append(i)
    return filename


def getTIDFileName(path, suffix):
    filename = []
    f_list = os.listdir(path)
    for i in f_list:
        if suffix.find(os.path.splitext(i)[1]) != -1:
            filename.append(i[1:3])
    return filename


def pil_loader(path):
    with open(path, 'rb') as f:
        img = Image.open(f)
        return img.convert('RGB')